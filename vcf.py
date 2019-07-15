from openpyxl import load_workbook
from openpyxl import Workbook
import os.path

def ask_user():
    check = str(input("Reply in (Y/N): ")).lower().strip()
    try:
        if check[0] == 'y':
            return True
        elif check[0] == 'n':
            return False
        else:
            print('Invalid Input')
            return ask_user()
    except Exception as error:
        print("Please enter valid inputs")
        print(error)
        return ask_user()


def vcfcreator(filename):                                   #main function
    if not os.path.isfile(filename):                        #checking if the file exists
        print("File doesn't exist. :(")
        return 0
    else:
        print("File found. :)\nProcessing...")
        book  = load_workbook(filename)
        print(book.sheetnames)                              #loading the file
        sheet = book.active                                 #loading the sheet
        file  = open("Contacts-llr.vcf","w+")               #creating the vcf file
        
        print(f"{sheet.cell(row=2, column=1).value} -- Is this the field containing the name?")
        print(f"{(sheet.cell(row=2, column=2).value)} -- and this the field containing the room number?")
        print(f"{int(sheet.cell(row=2, column=3).value)} -- and this the field containing the phone number?")
        
        if(ask_user() == True):
            with open("sample.vcf", "w+") as file:
                print(sheet.max_row)
                for i in range(2, sheet.max_row+1):
                    print(i)
                    if(sheet.cell(row=i, column=1).value == None or sheet.cell(row=i, column=2).value == None or sheet.cell(row=i, column=3).value == None):
                        print("One of the cell is empty.Check the file and try again  \nAborting...")
                        return 0
                    else:
                        file.write("BEGIN:VCARD\n")
                        file.write("VERSION:3.0\n")
                        file.write(f"FN:{sheet.cell(row=i, column=1).value} {sheet.cell(row=i, column=2).value} LLR\n")
                        file.write(f"N:{sheet.cell(row=i, column=1).value} {sheet.cell(row=i, column=2).value} LLR\n")
                        file.write(f"TEL;TYPE=CELL:{int(sheet.cell(row=i, column=3).value)}\n")
                        file.write(f"EMAIL:{(sheet.cell(row=i, column=4).value)}\n")
                        file.write("END:VCARD\n")
			
            print("Write Successful.\nThanks for using !!")		
        
        else:
            print("Correct the file format and try again.")
	
vcfcreator("sample.csv")